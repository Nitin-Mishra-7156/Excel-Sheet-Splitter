import openpyxl
import os

def split_excel_sheets(file_path, output_folder):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Create output directory if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Create a new workbook for each sheet
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name

        # Copy the data from the original sheet to the new workbook
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # Save the new workbook with the sheet name
        output_file = os.path.join(output_folder, f"{sheet_name}.xlsx")
        new_workbook.save(output_file)
        print(f"Sheet '{sheet_name}' saved to '{output_file}'.")

# Input and Output paths
input_file = "Your_file.xlsx"  # Replace with your input Excel file
output_directory = "output_sheets"  # Replace with your desired output folder

# Call the function
split_excel_sheets(input_file, output_directory)
